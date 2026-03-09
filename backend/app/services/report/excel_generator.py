"""Excel export for LLM scoring results.

Generates a single-sheet workbook with one row per assessable requirement,
grouped by framework with section-header rows separating each group.

Column layout (design depth):
  [Framework] [L1 Code] [L1 Name] ... [Req Code] [Req Name] [Description]
  [Status] [Skip Reason]
  [Score 1 (%)] [S1 Summary] [S1 Supporting Documents] [S1 Deficiencies] [S1 Improvements]
  [Score 2 (%)] [S2 Summary] [S2 Supporting Documents] [S2 Gap Analysis] [S2 Recommendations]
  [Score 3 (%)] [S3 Summary] [S3 Supporting Documents] [S3 Peer Analysis] [S3 Guidance]

Column layout (implementation depth — Score 1 split into sub-scores):
  [Framework] [L1 Code] [L1 Name] ... [Req Code] [Req Name] [Description]
  [Status] [Skip Reason]
  [Score 1 — Composite (%)]
  [Score 1 — Design (%)] [Design Summary] [Design Supporting Documents] [Design Deficiencies] [Design Improvements]
  [Score 1 — Implementation (%)] [Impl Summary] [Impl Supporting Documents] [Impl Deficiencies] [Impl Improvements]
  [Score 2 (%)] [S2 Summary] [S2 Supporting Documents] [S2 Gap Analysis] [S2 Recommendations]
  [Score 3 (%)] [S3 Summary] [S3 Supporting Documents] [S3 Peer Analysis] [S3 Guidance]
"""

import io
import uuid
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.assessment import Assessment
from app.models.scoring_job import RequirementScore
from app.models.unified_framework import (
    AssessmentFrameworkScope,
    Framework,
    FrameworkRequirement,
)

# ─── Colour palette ───────────────────────────────────────────────────────────

_HEADER_BG  = "2D4A6B"   # dark navy — main column header
_SECTION_BG = "E8F0FE"   # light blue — per-framework section banner

_GREEN  = "E6F4EA"        # score ≥ 75
_AMBER  = "FFF8E1"        # 50–74
_ORANGE = "FFF3E0"        # 25–49
_RED    = "FFEBEE"        # < 25
_SKIP   = "F5F5F5"        # N/A / skipped / not scored


def _score_fill(value: float | None) -> PatternFill:
    if value is None:
        return PatternFill(fill_type="solid", fgColor=_SKIP)
    if value >= 75:
        return PatternFill(fill_type="solid", fgColor=_GREEN)
    if value >= 50:
        return PatternFill(fill_type="solid", fgColor=_AMBER)
    if value >= 25:
        return PatternFill(fill_type="solid", fgColor=_ORANGE)
    return PatternFill(fill_type="solid", fgColor=_RED)


# ─── Explanation text helpers ─────────────────────────────────────────────────

def _fmt_items(items: list) -> str:
    """Flatten a list of explanation sub-items into bullet-separated text."""
    if not items:
        return ""
    lines: list[str] = []
    for item in items:
        if isinstance(item, str):
            lines.append(f"\u2022 {item}")
            continue
        if not isinstance(item, dict):
            continue
        # ID prefix (element_id or mechanism_id)
        id_part = item.get("element_id") or item.get("mechanism_id") or ""
        # Main text — try specific keys first, then level+action pattern, then fallback
        text = (
            item.get("issue")
            or item.get("gap")
            or item.get("recommendation")
            or item.get("guidance")
        )
        if not text:
            # Supporting document pattern: {title, relevant_details}
            if "title" in item:
                title = item.get("title", "")
                details = item.get("relevant_details", "")
                lines.append(f"\u2022 {title}: {details}" if details else f"\u2022 {title}")
                continue
            level = item.get("level")
            action = item.get("action")
            if level and action:
                text = f"[{level}] {action}"
            elif action:
                text = action
            else:
                text = " | ".join(str(v) for v in item.values() if v)
        prefix = f"{id_part}: " if id_part and text else ""
        lines.append(f"\u2022 {prefix}{text or ''}")
    return "\n".join(lines)


def _fmt_explanation(expl: dict | None, sublist_keys: list[str]) -> list[str]:
    """Return [summary, list1_text, list2_text, ...] from a score explanation dict.

    Always returns exactly (1 + len(sublist_keys)) strings.
    """
    if not expl:
        return [""] * (1 + len(sublist_keys))
    results = [expl.get("executive_summary") or ""]
    for key in sublist_keys:
        val = expl.get(key)
        results.append(_fmt_items(val) if isinstance(val, list) else (str(val) if val else ""))
    return results


# ─── Hierarchy traversal ──────────────────────────────────────────────────────

def _sort_by_hierarchy(
    assessable_reqs: list[FrameworkRequirement],
    req_map: dict[uuid.UUID, FrameworkRequirement],
    framework_id: uuid.UUID,
) -> list[FrameworkRequirement]:
    """Return assessable requirements in DFS pre-order following display_order."""
    children: dict[uuid.UUID | None, list[FrameworkRequirement]] = {}
    for r in req_map.values():
        if r.framework_id != framework_id:
            continue
        pid = r.parent_id
        if pid not in children:
            children[pid] = []
        children[pid].append(r)

    for pid in children:
        children[pid].sort(key=lambda r: (r.display_order or 0, r.code or ""))

    assessable_ids = {r.id for r in assessable_reqs}
    result: list[FrameworkRequirement] = []

    def _dfs(parent_id: uuid.UUID | None) -> None:
        for child in children.get(parent_id, []):
            if child.id in assessable_ids:
                result.append(child)
            _dfs(child.id)

    _dfs(None)
    return result


def _build_req_map(db: Session, framework_ids: list[uuid.UUID]) -> dict[uuid.UUID, FrameworkRequirement]:
    rows = (
        db.query(FrameworkRequirement)
        .filter(FrameworkRequirement.framework_id.in_(framework_ids))
        .all()
    )
    return {r.id: r for r in rows}


def _get_ancestors(req_id: uuid.UUID, req_map: dict[uuid.UUID, FrameworkRequirement]) -> list[FrameworkRequirement]:
    """Return [root, …, direct_parent] for a requirement using the in-memory map."""
    ancestors: list[FrameworkRequirement] = []
    current = req_map.get(req_id)
    while current and current.parent_id:
        parent = req_map.get(current.parent_id)
        if not parent:
            break
        ancestors.insert(0, parent)
        current = parent
    return ancestors


# ─── Main export ──────────────────────────────────────────────────────────────

def generate_scoring_export(db: Session, assessment_id: uuid.UUID) -> bytes:
    """Build and return raw XLSX bytes for the scoring export of one assessment."""

    assessment = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    if not assessment:
        raise ValueError(f"Assessment {assessment_id} not found")

    scope_rows = (
        db.query(AssessmentFrameworkScope)
        .filter(AssessmentFrameworkScope.assessment_id == assessment_id)
        .all()
    )
    if not scope_rows:
        raise ValueError("No frameworks in scope for this assessment")

    framework_ids = [s.framework_id for s in scope_rows]

    frameworks: dict[uuid.UUID, Framework] = {
        f.id: f
        for f in db.query(Framework).filter(Framework.id.in_(framework_ids)).all()
    }

    scores_by_req: dict[uuid.UUID, RequirementScore] = {
        s.requirement_id: s
        for s in db.query(RequirementScore)
        .filter(RequirementScore.assessment_id == assessment_id)
        .all()
    }

    req_map = _build_req_map(db, framework_ids)

    # Max number of ancestor columns needed across all frameworks
    max_anc = max((max(f.hierarchy_levels - 1, 0) for f in frameworks.values()), default=0)

    first_fw = next(iter(frameworks.values()), None)
    fw_labels: list[str] = list(first_fw.hierarchy_labels or []) if first_fw else []

    def _anc_label(i: int) -> str:
        return fw_labels[i] if i < len(fw_labels) - 1 else f"Level {i + 1}"

    def _req_label() -> str:
        return fw_labels[-1] if fw_labels else "Requirement"

    # ── Build header ──────────────────────────────────────────────────────────
    header: list[str] = ["Framework"]
    for i in range(max_anc):
        lbl = _anc_label(i)
        header += [f"{lbl} Code", f"{lbl} Name"]
    rl = _req_label()
    header += [f"{rl} Code", f"{rl} Name", "Description"]

    # Status / Skip Reason come before score columns
    header += ["Status", "Skip Reason"]

    # 0-based index of the first column after the hierarchy prefix
    # framework(1) + ancestors(2*N) + req code/name/desc(3) + status/skip(2)
    hier_width = 1 + max_anc * 2 + 3   # cols before status
    # status_col = hier_width (0-based)
    # score section starts at hier_width + 2

    score_start = hier_width + 2  # 0-based index of first score % column

    header += [
        "Score 1 — Coverage", "Score 1: Explanation", "Score 1: Recommendations",
    ]
    s1_col = score_start   # 0-based
    score1_width = 3

    s2_col = score_start + score1_width        # 0-based
    s3_col = s2_col + 5

    header += [
        "Score 2 (%)", "Score 2: Summary", "Score 2: Supporting Documents",
        "Score 2: Gap Analysis", "Score 2: Recommendations",
        "Score 3 (%)", "Score 3: Summary", "Score 3: Supporting Documents",
        "Score 3: Peer Analysis", "Score 3: Guidance",
    ]

    # ── Styles ────────────────────────────────────────────────────────────────
    FONT      = "Arial"
    hdr_font  = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    sec_font  = Font(name=FONT, bold=True, color="1A237E", size=10)
    data_font = Font(name=FONT, size=9)
    wrap_al   = Alignment(wrap_text=True, vertical="top")
    center_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_fill  = PatternFill(fill_type="solid", fgColor=_HEADER_BG)
    sec_fill  = PatternFill(fill_type="solid", fgColor=_SECTION_BG)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Scores"

    ws.append(header)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_al
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Data rows grouped by framework ────────────────────────────────────────
    for scope in scope_rows:
        fw = frameworks.get(scope.framework_id)
        if not fw:
            continue

        reqs = [r for r in req_map.values() if r.framework_id == fw.id and r.is_assessable]
        if not scope.include_all:
            included = {str(i) for i in (scope.included_requirement_ids or [])}
            reqs = [r for r in reqs if str(r.id) in included]
        else:
            excluded = {str(i) for i in (scope.excluded_requirement_ids or [])}
            reqs = [r for r in reqs if str(r.id) not in excluded]
        if not reqs:
            continue
        reqs = _sort_by_hierarchy(reqs, req_map, fw.id)

        # Section banner
        labels_here = list(fw.hierarchy_labels or [])
        hier_path = " \u2192 ".join(labels_here) if labels_here else f"{fw.hierarchy_levels} levels"
        banner = f"{fw.name}  \u2502  {hier_path}  \u2502  {len(reqs)} requirements"
        ws.append([banner] + [""] * (len(header) - 1))
        rn = ws.max_row
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=len(header))
        bc = ws.cell(row=rn, column=1)
        bc.font = sec_font
        bc.fill = sec_fill
        bc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[rn].height = 20

        # One row per requirement
        for req in reqs:
            ancestors = _get_ancestors(req.id, req_map)
            score = scores_by_req.get(req.id)

            # ── Hierarchy prefix ──
            row: list[Any] = [fw.name]
            for i in range(max_anc):
                if i < len(ancestors):
                    row += [ancestors[i].code or "", ancestors[i].name or ""]
                else:
                    row += ["", ""]
            row += [req.code or "", req.name or "", req.description or ""]

            # ── Status / Skip Reason (always filled) ──
            if score and score.status == "completed":
                row += ["completed", ""]
            elif score and score.status == "skipped":
                row += ["skipped", score.skip_reason or ""]
            elif score and score.status == "failed":
                row += ["failed", score.error_message or ""]
            else:
                row += ["not scored", ""]

            # ── Score columns ──
            if score and score.status == "completed":
                p5 = score.phase5_output or {}
                s1_expl = score.score1_explanation or {}
                coverage_label = s1_expl.get("coverage_label", "") if isinstance(s1_expl, dict) else ""
                explanation_text = s1_expl.get("explanation", "") if isinstance(s1_expl, dict) else ""
                recommendations = s1_expl.get("recommendations", []) if isinstance(s1_expl, dict) else []
                rec_text = "; ".join(
                    r.get("action", "") for r in recommendations if isinstance(r, dict) and r.get("action")
                )
                row += [coverage_label, explanation_text, rec_text]

                s2_expl = score.score2_explanation or p5.get("score2_explanation")
                s3_expl = score.score3_explanation or p5.get("score3_explanation")
                s2 = _fmt_explanation(s2_expl, ["supporting_documents", "gap_analysis", "recommendations"])
                s3 = _fmt_explanation(s3_expl, ["supporting_documents", "peer_analysis", "guidance"])
                row += [
                    round(score.score2) if score.score2 is not None else None,
                    s2[0], s2[1], s2[2], s2[3],
                    round(score.score3) if score.score3 is not None else None,
                    s3[0], s3[1], s3[2], s3[3],
                ]
            else:
                # All score cells blank
                row += [""] * (score1_width + 10)

            ws.append(row)
            dn = ws.max_row
            for cell in ws[dn]:
                cell.font = data_font
                cell.alignment = wrap_al

            # Colour-code score % cells (openpyxl columns are 1-based, so +1)
            score_pct_cols = [s2_col, s3_col]
            for ci in score_pct_cols:
                cell = ws.cell(row=dn, column=ci + 1)
                if isinstance(cell.value, (int, float)):
                    cell.fill = _score_fill(cell.value)
                    cell.alignment = center_al
                else:
                    cell.fill = _score_fill(None)
                    cell.alignment = center_al

    # ── Column widths ─────────────────────────────────────────────────────────
    widths: list[int] = [22]                       # Framework
    for _ in range(max_anc):
        widths += [10, 28]                         # ancestor code + name
    widths += [14, 50, 65]                         # req code, req name, description
    widths += [12, 28]                             # Status, Skip Reason

    widths += [14, 70, 70]                              # Score 1 (coverage label, explanation, recommendations)

    widths += [9, 55, 55, 55, 55]                       # Score 2
    widths += [9, 55, 55, 55, 55]                       # Score 3

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
